#!/usr/bin/env python3
import sqlite3
from pathlib import Path
from openpyxl import Workbook
from openpyxl.styles import (Font, PatternFill, Alignment, Border, Side,
                              GradientFill)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from openpyxl.chart.series import DataPoint
from openpyxl.chart import DoughnutChart
from openpyxl.formatting.rule import DataBarRule, ColorScaleRule

DB = Path(__file__).parent / "holdings.db"
OUT = Path(__file__).parent / "holdings.xlsx"

# ── helpers ──────────────────────────────────────────────────────────────────
BLUE   = Font(color="000000FF", name="Arial")
BLACK  = Font(color="FF000000", name="Arial")
HEADER_FILL = PatternFill("solid", fgColor="1A1A2E")
ALT_FILL    = PatternFill("solid", fgColor="F7F9FF")
ACCENT_FILL = PatternFill("solid", fgColor="EEF2FF")
UP_FILL     = PatternFill("solid", fgColor="FEE2E2")   # 上漲 → 紅
DN_FILL     = PatternFill("solid", fgColor="D1FAE5")   # 下跌 → 綠
NEW_FILL    = PatternFill("solid", fgColor="FFF1F2")   # 新增 → 淡紅底
DEL_FILL    = PatternFill("solid", fgColor="ECFDF5")   # 移除 → 淡綠底

def header_font(): return Font(bold=True, color="FFFFFFFF", name="Arial", size=10)
def title_font():  return Font(bold=True, color="FF1A1A2E", name="Arial", size=11)
def bold():        return Font(bold=True, name="Arial", size=9)
def normal():      return Font(name="Arial", size=9)
def blue_bold():   return Font(bold=True, color="000000FF", name="Arial", size=9)
def thin_border():
    s = Side(style="thin", color="E2E8F0")
    return Border(left=s, right=s, top=s, bottom=s)
def center(): return Alignment(horizontal="center", vertical="center")
def right():  return Alignment(horizontal="right",  vertical="center")
def left():   return Alignment(horizontal="left",   vertical="center")

def set_headers(ws, cols):
    for c, (hdr, w) in enumerate(cols, 1):
        cell = ws.cell(row=1, column=c, value=hdr)
        cell.font  = header_font()
        cell.fill  = HEADER_FILL
        cell.alignment = center()
        cell.border = thin_border()
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.row_dimensions[1].height = 22

def style_data_cell(cell, align=None, alt=False):
    cell.font   = normal()
    cell.border = thin_border()
    cell.fill   = ALT_FILL if alt else PatternFill("solid", fgColor="FFFFFF")
    if align:   cell.alignment = align

# ── fetch data ────────────────────────────────────────────────────────────────
conn = sqlite3.connect(DB)
snaps = conn.execute("""
    SELECT id, data_date FROM snapshots
    WHERE id IN (SELECT MAX(id) FROM snapshots GROUP BY data_date)
    ORDER BY data_date
""").fetchall()

all_data = {}
for snap_id, date in snaps:
    rows = conn.execute(
        "SELECT ticker, name, weight, shares FROM holdings WHERE snapshot_id=? ORDER BY weight DESC",
        (snap_id,)
    ).fetchall()
    all_data[date] = [{"ticker": r[0] or "", "name": r[1], "weight": r[2] or 0, "shares": r[3] or 0}
                      for r in rows]
conn.close()

dates = list(all_data.keys())
wb = Workbook()
wb.remove(wb.active)   # remove default sheet

# ── Per-date sheets ───────────────────────────────────────────────────────────
for date in dates:
    holdings = all_data[date]
    ws = wb.create_sheet(title=date)
    ws.freeze_panes = "A2"

    cols = [("排名", 6), ("股票名稱", 16), ("代碼", 8),
            ("投資比例(%)", 13), ("持有股數", 14)]
    set_headers(ws, cols)

    for i, h in enumerate(holdings, 2):
        alt = (i % 2 == 0)
        cells = [
            ws.cell(i, 1, i - 1),
            ws.cell(i, 2, h["name"]),
            ws.cell(i, 3, h["ticker"]),
            ws.cell(i, 4, h["weight"] / 100),
            ws.cell(i, 5, h["shares"]),
        ]
        aligns = [center(), left(), center(), right(), right()]
        for c, al in zip(cells, aligns):
            style_data_cell(c, al, alt)
        cells[1].font = bold()
        cells[2].font = blue_bold()
        cells[3].number_format = "0.00%"
        cells[4].number_format = '#,##0'

    # Title row above header
    ws.insert_rows(1)
    ws.merge_cells("A1:E1")
    t = ws["A1"]
    t.value = f"ETF 00981A 主動統一台股增長 — 持股清單  ({date})"
    t.font  = Font(bold=True, name="Arial", size=13, color="FFFFFFFF")
    t.fill  = PatternFill("solid", fgColor="4361EE")
    t.alignment = center()
    ws.row_dimensions[1].height = 28

    # Conditional: color scale on weight column
    last_row = len(holdings) + 2
    ws.conditional_formatting.add(
        f"D3:D{last_row}",
        ColorScaleRule(
            start_type="min",  start_color="FFFFFF",
            end_type="max",    end_color="4361EE"
        )
    )

    # ── Summary section to the right ──────────────────
    sc = 7
    ws.column_dimensions[get_column_letter(sc)].width = 22
    ws.column_dimensions[get_column_letter(sc + 1)].width = 14

    summary = [
        ("統計項目", "數值"),
        ("持股支數",   len(holdings)),
        ("總投資比例", sum(h["weight"] for h in holdings) / 100),
        ("前 10 大集中度", sum(h["weight"] for h in holdings[:10]) / 100),
        ("最大持股", holdings[0]["name"]),
        ("最大比例", holdings[0]["weight"] / 100),
        ("最小比例", holdings[-1]["weight"] / 100),
    ]
    for ri, (k, v) in enumerate(summary, 2):
        lc = ws.cell(ri, sc, k)
        vc = ws.cell(ri, sc + 1, v)
        lc.font   = bold();  lc.border = thin_border(); lc.alignment = left()
        vc.border = thin_border(); vc.alignment = right(); vc.font = normal()
        if ri == 2:
            lc.fill = HEADER_FILL; lc.font = header_font()
            vc.fill = HEADER_FILL; vc.font = header_font(); vc.alignment = center()
        else:
            row_fill = ALT_FILL if ri % 2 == 0 else PatternFill("solid", fgColor="FFFFFF")
            lc.fill = row_fill
            vc.fill = row_fill
        if isinstance(v, float):
            vc.number_format = "0.00%"

    # ── Bar chart (top 15) ─────────────────────────────
    chart = BarChart()
    chart.type    = "bar"
    chart.title   = f"前 15 大持股比例 ({date})"
    chart.y_axis.title = "投資比例"
    chart.x_axis.title = "股票"
    chart.style  = 10
    chart.width  = 18
    chart.height = 12

    data_ref = Reference(ws, min_col=4, min_row=2, max_row=17)   # header+15 rows
    cats_ref = Reference(ws, min_col=2, min_row=3, max_row=17)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.series[0].graphicalProperties.solidFill = "4361EE"
    ws.add_chart(chart, f"{get_column_letter(sc)}10")

# ── Diff sheet ────────────────────────────────────────────────────────────────
if len(dates) >= 2:
    old_date, new_date = dates[-2], dates[-1]
    old_map = {h["name"]: h for h in all_data[old_date]}
    new_map = {h["name"]: h for h in all_data[new_date]}
    all_names = set(old_map) | set(new_map)

    rows_out = []
    for name in all_names:
        o = old_map.get(name)
        n = new_map.get(name)
        if   not o: status = "新增"
        elif not n: status = "移除"
        elif o["weight"] != n["weight"] or o["shares"] != n["shares"]: status = "變化"
        else: status = "持平"
        old_s = o["shares"] if o else 0
        new_s = n["shares"] if n else 0
        rows_out.append({
            "name":    name,
            "ticker":  (o or n)["ticker"],
            "status":  status,
            "old_w":   o["weight"] if o else 0,
            "new_w":   n["weight"] if n else 0,
            "delta_w": (n["weight"] if n else 0) - (o["weight"] if o else 0),
            "old_s":   old_s,
            "new_s":   new_s,
            "delta_s": new_s - old_s,
        })
    # sort: 新增 → 移除 → 變化(by abs delta) → 持平
    order = {"新增": 0, "移除": 1, "變化": 2, "持平": 3}
    rows_out.sort(key=lambda r: (order[r["status"]], -abs(r["delta_w"])))

    ws = wb.create_sheet(title="差異比較")

    diff_cols = [("狀態", 8), ("股票名稱", 16), ("代碼", 8),
                 (f"舊比例% ({old_date})", 16), (f"新比例% ({new_date})", 16),
                 ("比例差異", 12),
                 (f"舊股數 ({old_date})", 16), (f"新股數 ({new_date})", 16),
                 ("股數差異", 14)]
    set_headers(ws, diff_cols)
    ws.row_dimensions[1].height = 22

    # Insert title row above headers
    ws.insert_rows(1)
    ws.merge_cells("A1:I1")
    t = ws["A1"]
    t.value = f"持股差異比較：{old_date}  →  {new_date}"
    t.font  = Font(bold=True, name="Arial", size=13, color="FFFFFFFF")
    t.fill  = PatternFill("solid", fgColor="4361EE")
    t.alignment = center()
    ws.row_dimensions[1].height = 28
    ws.freeze_panes = "A3"

    STATUS_FILL = {"新增": NEW_FILL, "移除": DEL_FILL, "變化": ACCENT_FILL,
                   "持平": PatternFill("solid", fgColor="FFFFFF")}
    STATUS_FONT_COLOR = {"新增": "DC2626", "移除": "16A34A", "變化": "1E40AF", "持平": "555555"}

    for i, r in enumerate(rows_out, 3):
        alt = i % 2 == 0
        sf  = STATUS_FILL[r["status"]]
        sc  = STATUS_FONT_COLOR[r["status"]]

        def wc(col, val):
            c = ws.cell(i, col, val)
            c.border = thin_border()
            c.fill   = sf
            c.font   = Font(name="Arial", size=9)
            return c

        s = wc(1, r["status"]); s.font = Font(bold=True, name="Arial", size=9, color=sc); s.alignment = center()
        nm = wc(2, r["name"]);   nm.font = bold(); nm.alignment = left()
        tk = wc(3, r["ticker"]); tk.font = blue_bold(); tk.alignment = center()
        ow = wc(4, r["old_w"] / 100 if r["old_w"] else None); ow.number_format = "0.00%"; ow.alignment = right()
        nw = wc(5, r["new_w"] / 100 if r["new_w"] else None); nw.number_format = "0.00%"; nw.alignment = right()
        dw = wc(6, r["delta_w"] / 100 if r["delta_w"] else None)
        dw.number_format = '+0.00%;-0.00%;"-"'
        dw.alignment = right()
        if r["delta_w"] > 0:
            dw.font = Font(bold=True, color="DC2626", name="Arial", size=9)   # 增加 → 紅
        elif r["delta_w"] < 0:
            dw.font = Font(bold=True, color="16A34A", name="Arial", size=9)   # 減少 → 綠
        os_ = wc(7, r["old_s"] or None); os_.number_format = "#,##0"; os_.alignment = right()
        ns  = wc(8, r["new_s"] or None); ns.number_format = "#,##0"; ns.alignment = right()
        ds  = wc(9, r["delta_s"] if r["delta_s"] != 0 else None)
        ds.number_format = '+#,##0;-#,##0;"-"'; ds.alignment = right()
        if r["delta_s"] > 0:
            ds.font = Font(bold=True, color="DC2626", name="Arial", size=9)   # 增加 → 紅
        elif r["delta_s"] < 0:
            ds.font = Font(bold=True, color="16A34A", name="Arial", size=9)   # 減少 → 綠

    # ── Summary chips row ──────────────────────────────
    cnts = {s: sum(1 for r in rows_out if r["status"]==s) for s in ["新增","移除","變化","持平"]}
    sr = len(rows_out) + 5
    labels = [("新增持股", cnts["新增"], "DC2626", "FEE2E2"),   # 新增 → 紅
              ("移除持股", cnts["移除"], "16A34A", "D1FAE5"),   # 移除 → 綠
              ("比例變化", cnts["變化"], "1E40AF", "DBEAFE"),
              ("持平",     cnts["持平"], "555555", "F3F4F6")]
    for ci, (lbl, cnt, fc, bg) in enumerate(labels):
        c1 = ws.cell(sr,     ci*2+1, lbl)
        c2 = ws.cell(sr + 1, ci*2+1, cnt)
        ws.merge_cells(start_row=sr, start_column=ci*2+1, end_row=sr, end_column=ci*2+2)
        ws.merge_cells(start_row=sr+1, start_column=ci*2+1, end_row=sr+1, end_column=ci*2+2)
        c1.fill = PatternFill("solid", fgColor=bg); c1.font = Font(bold=True, name="Arial", size=9, color=fc); c1.alignment = center()
        c2.fill = PatternFill("solid", fgColor=bg); c2.font = Font(bold=True, name="Arial", size=14, color=fc); c2.alignment = center()
        c1.border = thin_border(); c2.border = thin_border()

# ── Summary overview sheet ────────────────────────────────────────────────────
ws_sum = wb.create_sheet(title="總覽", index=0)
ws_sum.column_dimensions["A"].width = 20
ws_sum.column_dimensions["B"].width = 14

ws_sum.merge_cells("A1:C1")
t = ws_sum["A1"]
t.value = "ETF 00981A 主動統一台股增長 — 快照總覽"
t.font  = Font(bold=True, name="Arial", size=14, color="FFFFFFFF")
t.fill  = PatternFill("solid", fgColor="1A1A2E")
t.alignment = center()
ws_sum.row_dimensions[1].height = 30

headers = [("資料日期", 14), ("持股支數", 12), ("前10大集中度", 16)]
for c, (h, w) in enumerate(headers, 1):
    cell = ws_sum.cell(2, c, h)
    cell.font = header_font(); cell.fill = HEADER_FILL
    cell.alignment = center(); cell.border = thin_border()
    ws_sum.column_dimensions[get_column_letter(c)].width = w

for i, date in enumerate(dates, 3):
    h_list = all_data[date]
    top10 = sum(h["weight"] for h in h_list[:10]) / 100
    alt = i % 2 == 0
    for c, v in enumerate([date, len(h_list), top10], 1):
        cell = ws_sum.cell(i, c, v)
        cell.font   = bold()
        cell.fill   = ALT_FILL if alt else PatternFill("solid", fgColor="FFFFFF")
        cell.border = thin_border()
        cell.alignment = center()
        if c == 3: cell.number_format = "0.00%"

wb.save(OUT)
print(f"Excel saved: {OUT}")
print(f"File size: {OUT.stat().st_size / 1024:.1f} KB")
print(f"Sheets: {wb.sheetnames}")
